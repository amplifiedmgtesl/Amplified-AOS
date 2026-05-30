"""Pre-cutover cleanup workbook — PROD data.

Replaces the earlier dev-data workbook. The plan is now: clean prod
FIRST, then apply the queued migrations against clean data. Saves us
from cleaning the same records twice (once on dev, once on prod after
clone-back).

Output: docs/cleanup-review-prod-2026-05-26.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "C:/amplified/Amplified-AOS/docs/cleanup-review-prod-2026-05-26.xlsx"

HEADER_FILL = PatternFill("solid", start_color="2A3F5A")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Calibri", bold=True, size=14)
NOTE_FONT   = Font(name="Calibri", italic=True, size=10, color="555555")
ID_FONT     = Font(name="Consolas", size=10, color="666666")
DECISION_FILL = PatternFill("solid", start_color="FFF7E0")
URGENT_FILL = PatternFill("solid", start_color="FBEAEA")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def write_headers(ws, headers, start_row=1):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[start_row].height = 30

def write_row(ws, row_idx, values, *, id_cols=None, decision_col=None, urgent=False):
    id_cols = id_cols or set()
    for col_idx, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=v)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = THIN_BORDER
        if col_idx in id_cols:
            c.font = ID_FONT
        if decision_col is not None and col_idx == decision_col:
            c.fill = DECISION_FILL
        elif urgent and col_idx == 1:
            c.fill = URGENT_FILL

# ─────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════ README ════════════════════════
ws = wb.create_sheet("README")
ws["A1"] = "Pre-cutover cleanup checklist — PROD data"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Date: 2026-05-26   |   Env: prod (Supabase wmssllfmahotppoyxxrr)"
ws["A3"].font = NOTE_FONT
ws["A5"] = "Why this workbook exists"
ws["A5"].font = Font(bold=True, size=12)
ws["A6"] = (
    "The Jobs/quote/invoice/timekeeping rewrite has ~38 migrations queued for "
    "prod (see memory: project_pending_prod_migrations.md). Some of those "
    "migrations do data backfills that work best on clean data. If we leave "
    "stragglers in the data, post-cutover we'll either:\n"
    "  • have FK NOT NULL constraints that fail to install, OR\n"
    "  • re-do the same cleanup we already did on dev, against prod.\n\n"
    "Better: clean prod first, then run migrations against clean data. After "
    "prod's cutover, we can re-clone prod → dev so both envs are aligned."
)
ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[6].height = 130

ws["A8"] = "How to use"
ws["A8"].font = Font(bold=True, size=12)
ws["A9"] = (
    "Each sheet lists affected records WITH THEIR PROD RECORD GUIDs. "
    "Fill the pale-yellow 'Decision (Connor)' column with: KEEP / DELETE / "
    "REPOINT TO <id> / MERGE WITH <id> / BACKFILL / NEEDS DISCUSSION. "
    "Add notes if useful. After Connor's pass, I'll generate the cleanup SQL "
    "and run it on prod BEFORE the migration batch.\n\n"
    "Note: prod doesn't have the 'Show IDs' sidebar toggle yet — that ships "
    "with the merge to main. For now, IDs live only in this workbook. URL "
    "patterns like /quotes/<id> and /invoices/<id> still let you spot-check "
    "specific records."
)
ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[9].height = 130

ws["A11"] = "Sheet index"
ws["A11"].font = Font(bold=True, size=12)
index_rows = [
    ("1. Timesheet stragglers",       "3 timesheets with no clean job_request match. Must resolve BEFORE migration #31."),
    ("2. Quote lines NULL FK",        "11 recovered-* quotes / 127 lines with NULL specialty_id. Connor PDF recovery records."),
    ("3. Invoice lines NULL FK",      "67 invoice lines with NULL position_id and/or specialty_id. Some are LIVE active work — possible bug."),
    ("4. Duplicate job_requests",     "Carolina, Bruno Mars, Revival Night, KY Event (x4!) — all need merge decisions."),
    ("5. Empty job_request rows",     "2 totally blank job_request rows. Almost certainly DELETE."),
    ("6. Jobs missing crew_needs",    "Active jobs with 0 crew planning rows. Mostly LEAVE (past)."),
    ("7. Auto-resolved (FYI)",        "Already clean on prod — code cleanup only, no decision needed."),
]
for i, (name, desc) in enumerate(index_rows, start=12):
    ws.cell(row=i, column=1, value=name).font = Font(bold=True)
    ws.cell(row=i, column=2, value=desc)
    ws.row_dimensions[i].height = 22

ws["A20"] = "Urgency callout"
ws["A20"].font = Font(bold=True, size=12, color="8A1A1A")
ws["A21"] = (
    "🔴 Sheet 4 — Carolina Country Music Fest event STARTS 2026-05-31 (5 days out). "
    "The two duplicate job_requests have split data: one has crew_needs (40 rows), "
    "the other has the quote (79 lines). Connor needs to decide on this BEFORE the "
    "event regardless of the migration cutover timeline.\n\n"
    "🔴 Sheet 3 — recent NULL-FK invoice lines suggest a LIVE bug where new invoice "
    "lines aren't getting position_id/specialty_id populated by the legacy invoice "
    "builder. Worth investigating today rather than after cutover."
)
ws["A21"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[21].height = 100

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 95

# ════════════════ Sheet 1: Timesheet stragglers ════════════════
ws = wb.create_sheet("1. Timesheet stragglers")
ws["A1"] = "Cleanup #1 — Timesheets with no job_request match"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Same 3 records that exist on dev (these predate the dev/prod split). "
            "Must resolve BEFORE migration #31 runs the Phase 1 backfill — otherwise these "
            "stay as legacy/NULL job_id forever. Decisions are independent of dev's cleanup.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 60
ws.merge_cells("A2:H2")

headers = [
    "Timesheet ID", "Job Sheet ID", "JS client", "JS event", "JS date",
    "Entries", "w/ hours", "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

rows_data = [
    ("timesheet-jobsheet-1774846347019", "jobsheet-1774846347019",
     "Sunbelt Ground Protection Division", "Flooring", "2026-03-28",
     4, 2,
     "DELETE — past, no quote, no approval, no invoice. Only 2/4 rows have hours.",
     "", ""),
    ("timesheet-jobsheet-1774846277546", "jobsheet-1774846277546",
     "Sunbelt Ground Protection DIvision (typo)", "Flooring Install", "2026-03-28",
     9, 0,
     "DELETE — empty skeleton; near-duplicate of above with typo.",
     "", ""),
    ("timesheet-jobsheet-1774844548762", "jobsheet-1774844548762",
     "Loud and Clear (sep. clients row from Loud&Clear, Inc.)",
     "Spring Concert", "2026-04-17",
     13, 13,
     "REPOINT TO jobreq-1775744267941 (Mount St. Joseph - Spring Concert). Same date + Loud&Clear venue type. Then merge/dedupe the 'loud and clear' client row.",
     "", ""),
]
for i, r in enumerate(rows_data, start=5):
    write_row(ws, i, r, id_cols={1, 2}, decision_col=9)
    ws.row_dimensions[i].height = 65

widths = [38, 32, 35, 30, 12, 9, 9, 60, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ════════════════ Sheet 2: Quote lines NULL FK ═══════════════════
ws = wb.create_sheet("2. Quote lines NULL FK")
ws["A1"] = "Cleanup #2 — recovered-* quotes with NULL specialty_id (127 lines)"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "All 127 affected lines belong to 11 frozen recovered-* quotes from the Connor "
    "PDF recovery (memory: project_pdf_data_recovery.md). They render correctly today "
    "via the legacy text fallback. Three options per quote:\n"
    "  • BACKFILL = name-match specialty_id on lines (totals unchanged). Lets us "
    "retire the fallback code and add NOT NULL.\n"
    "  • LEAVE = accept these as historical records that use the legacy display path. "
    "Means we keep some fallback code in place for these IDs.\n"
    "  • DELETE = drop the recovered quote (only if not needed for audit / reference)."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 130
ws.merge_cells("A2:J2")

headers = [
    "Quote ID", "Status", "Event name", "Client",
    "Lines w/ NULL spec", "Total $",
    "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

# Sorted by total $ desc to surface biggest impact first
quote_rows = sorted([
    ("recovered-06a99ec9-pro-football-hall-of-fame-2026-enshrinement-week",
     "signed", "Pro Football Hall of Fame 2026 Enshrinement Week",
     "FEP Live, LLC", 54, 66977,
     "BACKFILL — signed, sizable, future event (Aug 2026).", "", ""),
    ("recovered-a65a9683-2026-farm-tour-california",
     "quoted", "2026 farm Tour - California",
     "Row Crop LLC - Vaden Group", 9, 66300,
     "BACKFILL — sizable, recent.", "", ""),
    ("recovered-b12d4439-the-ohio-country-fest",
     "quoted", "The Ohio Country Fest",
     "Aaron Green - Jayson Entertainment Group", 11, 44592,
     "BACKFILL — sizable, near-future (June 2026).", "", ""),
    ("recovered-6e82573f-miami-university-commencement",
     "signed", "Miami University Commencement",
     "Loud&Clear, Inc.", 23, 36918,
     "BACKFILL — signed.", "", ""),
    ("recovered-74bb42d2-luke-combs-load-out",
     "quoted", "Luke Combs - Load Out", "Rhino Staging", 5, 7648,
     "BACKFILL.", "", ""),
    ("recovered-ba7ea741-scotty-mccreery",
     "signed", "Scotty Mccreery", "Chris Stewart - Morris Farms", 8, 6665,
     "BACKFILL — signed, recent. (Also see Sheet 3 — current invoice work on this client has NULL FKs.)", "", ""),
    ("recovered-3d131e98-warrior-conference",
     "quoted", "Warrior Conference", "Loud&Clear, Inc.", 4, 4849,
     "BACKFILL — OR DELETE one if this duplicates the row below.", "", ""),
    ("recovered-1c3e90df-warrior-conference",
     "quoted", "Warrior Conference", "Loud&Clear, Inc.", 4, 4849,
     "SAME totals as recovered-3d131e98 above — likely a duplicate. DELETE one?", "", ""),
    ("recovered-79529235-mount-st-joseph-spring-concert",
     "signed", "Mount St. Joseph - Spring Concert", "Loud&Clear, Inc.", 6, 4480,
     "BACKFILL — signed.", "", ""),
    ("recovered-20871778-osu-stadium-load-out",
     "quoted", "OSU Stadium - Load Out", "Rhino Staging", 2, 3664,
     "BACKFILL.", "", ""),
    ("recovered-33c5d8eb-ky-event",
     "quoted", "KY Event", "Loud&Clear, Inc.", 1, 740,
     "BACKFILL (or DELETE if the underlying KY Event chain is being collapsed — see Sheet 4).", "", ""),
], key=lambda r: -r[5])

for i, r in enumerate(quote_rows, start=5):
    write_row(ws, i, r, id_cols={1}, decision_col=8)
    ws.row_dimensions[i].height = 55

widths = [55, 10, 50, 35, 11, 12, 65, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ════════════════ Sheet 3: Invoice lines NULL FK ════════════════
ws = wb.create_sheet("3. Invoice lines NULL FK")
ws["A1"] = "Cleanup #3 — invoice_lines with NULL position_id / specialty_id (67 lines)"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "Mix of historical recovered-* records AND recent live invoice work — the recent "
    "ones (highlighted, marked LIVE) suggest the legacy invoice-builder isn't populating "
    "position_id/specialty_id on new lines. Worth investigating today.\n"
    "Most rows have legacy 'department' / 'specialty' text that drives a clean "
    "name-match BACKFILL. Junk rows (empty department, zero hours) → DELETE."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 75
ws.merge_cells("A2:M2")

headers = [
    "Line ID", "Invoice ID", "Invoice no", "Status",
    "Department (legacy)", "Specialty (legacy)",
    "Hours", "Qty", "Total $",
    "Era", "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

# Group by source (RECENT live work vs HISTORICAL recovered-*)
invoice_rows = [
    # ── HISTORICAL recovered batch (these match the original recovery period)
    ("inv-recovered-58d712e5-line-1", "inv-recovered-58d712e5", "INV-2026-0330-246",
     "sent", "Labor", "", 10, 2, 740,
     "historical", "BACKFILL — 'Labor' → pos-01 Stagehand.", "", ""),
    ("inv-recovered-ff46be24-line-1", "inv-recovered-ff46be24", "INV-2026-0330-627",
     "sent", "Labor", "", 5, 3, 540,
     "historical", "BACKFILL — 'Labor' → pos-01.", "", ""),
    ("inv-recovered-e6894771-line-1", "inv-recovered-e6894771", "INV-2026-0410-503",
     "draft", "Stagehand", "", 2.5, 12, 1050,
     "historical", "BACKFILL — pos-01.", "", ""),
    ("inv-recovered-e6894771-line-2", "inv-recovered-e6894771", "INV-2026-0410-503",
     "draft", "Stagehand", "", 2.5, 12, 1050,
     "historical", "BACKFILL — note: identical to line-1. Verify not a dupe.", "", ""),
    ("inv-recovered-c8ecc32d-line-1", "inv-recovered-c8ecc32d", "INV-2026-0419-808",
     "sent", "Operations", "", 7.5, 1, 300,
     "historical", "BACKFILL — 'Operations' → pos+spc needs Connor's mapping (Crew Chief? Coordinator?).", "", ""),
    ("inv-recovered-770ff8c3-line-1", "inv-recovered-770ff8c3", "INV-2026-0419-808",
     "sent", "Operations", "", 7.5, 1, 300,
     "historical", "BACKFILL — twin of above (both 'sent', same invoice_no, may be a paired draft).", "", ""),
    ("inv-recovered-770ff8c3-line-2", "inv-recovered-770ff8c3", "INV-2026-0419-808",
     "sent", "Stagehand", "", 7.5, 11, 3052.5,
     "historical", "BACKFILL — pos-01.", "", ""),
    ("inv-recovered-c8ecc32d-line-2", "inv-recovered-c8ecc32d", "INV-2026-0419-808",
     "sent", "Stagehand", "", 7.5, 11, 3052.5,
     "historical", "BACKFILL — pos-01.", "", ""),
    ("inv-recovered-770ff8c3-line-3", "inv-recovered-770ff8c3", "INV-2026-0419-808",
     "sent", "Operations", "", 5, 1, 200, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-c8ecc32d-line-3", "inv-recovered-c8ecc32d", "INV-2026-0419-808",
     "sent", "Operations", "", 5, 1, 200, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-c8ecc32d-line-4", "inv-recovered-c8ecc32d", "INV-2026-0419-808",
     "sent", "Fork Op", "", 5, 1, 190, "historical", "BACKFILL — 'Fork Op' → pos-08 Forklift Operator.", "", ""),
    ("inv-recovered-770ff8c3-line-4", "inv-recovered-770ff8c3", "INV-2026-0419-808",
     "sent", "Fork Op", "", 5, 1, 190, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-c8ecc32d-line-5", "inv-recovered-c8ecc32d", "INV-2026-0419-808",
     "sent", "Stagehand", "", 5, 10, 1850, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-770ff8c3-line-5", "inv-recovered-770ff8c3", "INV-2026-0419-808",
     "sent", "Stagehand", "", 5, 10, 1850, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-922d482c-line-1", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Operations", "Crew Chief", 15, 1, 600,
     "historical", "BACKFILL — has both legacy fields; name-match should work cleanly.", "", ""),
    ("inv-recovered-922d482c-line-2", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Stagehand", "Labor", 6, 4, 912, "historical", "BACKFILL — pos-01 + spc-01-01.", "", ""),
    ("inv-recovered-922d482c-line-3", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Stagehand", "Labor", 7, 8, 2128, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-922d482c-line-4", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Stagehand", "Labor", 8, 2, 608, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-922d482c-line-5", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Forklift Operator", "Shop", 15, 1, 600,
     "historical", "BACKFILL — pos-08 + spc-08-01.", "", ""),
    ("inv-recovered-922d482c-line-6", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Stagehand", "Labor", 15, 5, 2850, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-922d482c-line-7", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Stagehand", "Labor", 8, 2, 608, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-922d482c-line-8", "inv-recovered-922d482c", "INV-2026-0424-352",
     "partial", "Stagehand", "Labor", 9, 1, 342, "historical", "BACKFILL.", "", ""),
    ("inv-1777044736028_2", "inv-1777044736028", "INV-2026-0424-352",
     "draft", "(empty)", "(empty)", 0, 1, 0, "junk", "DELETE — empty department/specialty, zero hours, zero total.", "", ""),
    ("inv-1777044736028_3", "inv-1777044736028", "INV-2026-0424-352",
     "draft", "(empty)", "(empty)", 0, 1, 0, "junk", "DELETE — junk row.", "", ""),
    ("inv-recovered-9601b91a-line-1", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Lead", "", 14.5, 1, 580, "historical", "BACKFILL — 'Lead' → pos-02 Stagehand Lead.", "", ""),
    ("inv-recovered-4a0b36ed-line-1", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Lead", "", 13.5, 1, 540, "historical", "BACKFILL — pos-02. (Note: PAST DUE status.)", "", ""),
    ("inv-recovered-4a0b36ed-line-2", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Stagehands", "", 10, 11, 4070, "historical", "BACKFILL — pos-01.", "", ""),
    ("inv-recovered-9601b91a-line-2", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Stagehands", "", 10, 11, 4070, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-3", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Stagehands", "", 13.5, 10, 4995, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-3", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Stagehands", "", 14.5, 10, 5365, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-4", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Working Fork", "", 13.5, 1, 513,
     "historical", "BACKFILL — 'Working Fork' → pos-08 Forklift Operator. Specialty?", "", ""),
    ("inv-recovered-9601b91a-line-4", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Working Fork", "", 14.5, 1, 551, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-5", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Meal Penalty", "", 3.5, 12, 1470,
     "historical", "'Meal Penalty' isn't a position. LEAVE or recategorize as a flat-fee/penalty line type.", "", ""),
    ("inv-recovered-4a0b36ed-line-5", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Meal Penalty", "", 3.5, 12, 1470, "historical", "Same as above.", "", ""),
    ("inv-recovered-9601b91a-line-6", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Lead", "", 5.5, 1, 220, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-6", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Lead", "", 5.5, 1, 220, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-7", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Stagehands", "", 5.5, 19, 3866.5, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-7", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Stagehands", "", 5.5, 19, 3866.5, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-8", "inv-recovered-4a0b36ed", "LC260311",
     "PAST DUE", "Working Fork", "", 5.5, 1, 209, "historical", "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-8", "inv-recovered-9601b91a", "LC260311",
     "DUE", "Working Fork", "", 5.5, 1, 209, "historical", "BACKFILL.", "", ""),
    # ── LIVE / RECENT — possible writer bug
    ("inv-1777933898166-recovered-1c3e90df-warrior-conference_0",
     "inv-1777933898166-recovered-1c3e90df-warrior-conference", "INV-2026-0504-819",
     "draft", "Operations", "Crew Chief", 7.5, 1, 300,
     "🔴 LIVE", "BACKFILL — but also INVESTIGATE: why is the live invoice-builder creating new lines with NULL FKs?", "", ""),
    ("inv-1777933898166-recovered-1c3e90df-warrior-conference_1",
     "inv-1777933898166-recovered-1c3e90df-warrior-conference", "INV-2026-0504-819",
     "draft", "Stagehand", "Labor", 7.5, 8, 2220,
     "🔴 LIVE", "Same — live writer bug suspect.", "", ""),
    ("inv-1779804866522_79", "inv-1779804866522", "INV-2026-0526-455",
     "draft", "(empty)", "(empty)", 0, 1, 0,
     "🔴 LIVE", "DELETE — junk live row.", "", ""),
    # Scotty McCreery cluster — multiple drafts in flight
    ("inv-1779818311948_0", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Operations", "Crew Chief", 6, 1, 240,
     "🔴 LIVE", "BACKFILL — Scotty McCreery, sent. Live writer bug.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_0",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 5, 1, 200,
     "🔴 LIVE", "BACKFILL — same event.", "", ""),
    ("inv-1779819449197_0", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 6, 1, 240,
     "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_1", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Stagehand", "Labor", 6, 5, 1050, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_1",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 5, 5, 875, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_1", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 6, 5, 1050, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_2",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 5, 1, 200, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_2", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 5, 1, 200, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_2", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Operations", "Crew Chief", 5, 1, 200, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_3", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 5, 5, 875, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_3",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 5, 5, 875, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_3", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Stagehand", "Labor", 5, 5, 875, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_4", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 7, 1, 280, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_4",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 16, 1, 640, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_4", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Operations", "Crew Chief", 7, 1, 280, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_5", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 7, 5, 1225, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_5", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Stagehand", "Labor", 7, 5, 1225, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_5",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 16, 5, 2800, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_6",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 5, 1, 200, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_6", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Operations", "Crew Chief", 8, 1, 320, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_6", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Operations", "Crew Chief", 8, 1, 320, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818021784-recovered-ba7ea741-scotty-mccreery_7",
     "inv-1779818021784-recovered-ba7ea741-scotty-mccreery", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 5, 5, 875, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779819449197_7", "inv-1779819449197", "INV-2026-0526-934 AES_26052124_MOR",
     "draft", "Stagehand", "Labor", 8, 5, 1400, "🔴 LIVE", "BACKFILL.", "", ""),
    ("inv-1779818311948_7", "inv-1779818311948", "INV-2026-0526-934 AES_26052124_MOR",
     "sent", "Stagehand", "Labor", 8, 5, 1400, "🔴 LIVE", "BACKFILL.", "", ""),
]
for i, r in enumerate(invoice_rows, start=5):
    is_live = "LIVE" in str(r[9])
    write_row(ws, i, r, id_cols={1, 2}, decision_col=12, urgent=is_live)
    ws.row_dimensions[i].height = 38

widths = [42, 38, 25, 12, 22, 18, 8, 7, 11, 11, 50, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ════════════════ Sheet 4: Duplicate job_requests ════════════════
ws = wb.create_sheet("4. Duplicate job_requests")
ws["A1"] = "Cleanup #4 — duplicate job_request clusters"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "Multiple job_request rows for the same real-world event. Free-text intake "
    "with no dedupe at the Job Request screen is the root cause (a separate intake-"
    "duplicate-warning feature is queued). Resolve by picking the canonical row and "
    "marking the others SUPERSEDE or DELETE.\n\n"
    "🔴 Carolina Country Music Fest is most urgent — event starts 2026-05-31."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 90
ws.merge_cells("A2:K2")

headers = [
    "Cluster", "Job request ID", "Status", "Request date", "End date",
    "Days", "Crew needs", "Quotes", "Linked attachments / quote",
    "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

dupe_rows = [
    # ── Carolina cluster
    ("🔴 Carolina Country Music Fest (URGENT — starts 2026-05-31)",
     "jobreq-1778348194976", "lead", "2026-05-31", "2026-06-10", 10, 0, 1,
     "Has quote with 79 lines; 0 crew_needs",
     "MERGE — move the quote (or its line content) onto jobreq-1779670159567, then DELETE this row.", "", ""),
    ("🔴 Carolina Country Music Fest (URGENT)",
     "jobreq-1779670159567", "lead", "2026-05-31", "2026-06-09", 10, 40, 0,
     "Has crew_needs; 1 attachment; no quote",
     "KEEP as canonical — has the planning + attachment.", "", ""),
    # ── Bruno Mars cluster
    ("Bruno Mars (past — 2026-05-16 → 21)",
     "jobreq-1778094212255", "lead", "2026-05-16", "2026-05-21", 5, 12, 1,
     "Fuller record — 5 days, 12 crew_needs, 1 quote",
     "KEEP as canonical.", "", ""),
    ("Bruno Mars",
     "jobreq-1777684960205", "lead", "2026-05-16", "2026-05-21", 4, 6, 0,
     "Half-baked — 4 days, 6 crew_needs, no quote",
     "SUPERSEDE/DELETE — abandoned duplicate.", "", ""),
    # ── Revival Night cluster (memory entry #2 already flagged this)
    ("Revival Night (already in pending #2)",
     "jobreq-1775344443515", "lead", "2026-04-17", "2026-04-17", 1, 0, 0,
     "Has real notes content",
     "KEEP per memory entry #2.", "", ""),
    ("Revival Night",
     "jobreq-1775227265513", "lead", "2026-04-17", "2026-04-17", 1, 0, 0,
     "Empty sibling",
     "DELETE per memory entry #2.", "", ""),
    # ── KY Event cluster — FOUR rows on prod, not three
    ("KY Event (was 3 dupes in memory #2 — actually 4 on prod)",
     "jobreq-1775346228492", "booked", "2026-04-05", "2026-04-05", 1, 0, 0,
     "Memory #2 said keep this one (has linked quote)",
     "KEEP per memory entry #2.", "", ""),
    ("KY Event",
     "jobreq-1775346126232", "booked", "2026-04-05", "2026-04-05", 1, 0, 0, "",
     "DELETE per memory entry #2.", "", ""),
    ("KY Event",
     "jobreq-1775345942610", "booked", "2026-04-05", "2026-04-05", 1, 0, 0, "",
     "DELETE per memory entry #2.", "", ""),
    ("KY Event — NEW 4th row not in memory #2",
     "jobreq-1775064575180", "lead", "2026-03-31", "2026-03-31", 1, 0, 0,
     "Different date (2026-03-31 vs others on 2026-04-05). Could be a separate event or another dupe.",
     "NEEDS DISCUSSION — same client + similar event_name, different date. Separate event or 4th dupe?", "", ""),
]
for i, r in enumerate(dupe_rows, start=5):
    is_urgent = "🔴" in str(r[0])
    write_row(ws, i, r, id_cols={2}, decision_col=11, urgent=is_urgent)
    ws.row_dimensions[i].height = 55

widths = [40, 32, 10, 13, 13, 7, 11, 8, 35, 55, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ════════════════ Sheet 5: Empty job_requests ════════════════
ws = wb.create_sheet("5. Empty job_request rows")
ws["A1"] = "Cleanup #5 — totally blank job_request rows"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("2 job_request rows on prod with all human-facing fields blank "
            "(empty event_name, client, dates). Almost certainly accidental rows "
            "that never got fleshed out. Verify no downstream references then DELETE.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 50
ws.merge_cells("A2:F2")

headers = ["Job request ID", "Status", "Has any data?", "Recommendation",
           "Decision (Connor)", "Notes"]
write_headers(ws, headers, start_row=4)

empty_rows = [
    ("jobreq-1775064576002", "lead", "No — event_name='', client='', no dates",
     "DELETE after verifying no quote/invoice/attachment references it.", "", ""),
    ("jobreq-1776229712651", "lead", "No — same",
     "DELETE after verification.", "", ""),
]
for i, r in enumerate(empty_rows, start=5):
    write_row(ws, i, r, id_cols={1}, decision_col=5)
    ws.row_dimensions[i].height = 45

widths = [32, 10, 50, 55, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ════════════════ Sheet 6: Jobs missing crew_needs ════════════════
ws = wb.create_sheet("6. Jobs missing crew_needs")
ws["A1"] = "Cleanup #6 — active job_requests with 0 crew_needs"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "Active job_requests (status not 'completed'/'cancelled') with no per-day crew planning rows.\n"
    "Per Connor's earlier guidance: PAST jobs can stay empty (work is done; crew planning is moot). "
    "FUTURE jobs need crew_needs populated — drives the planning + invoice flow.\n"
    "Most rows here are past. Highlighting future ones for action."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 75
ws.merge_cells("A2:I2")

headers = [
    "Job request ID", "Event", "Client", "Status",
    "Start", "End", "Days",
    "Era", "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

# Today is 2026-05-26 per the system prompt
# Future = start >= 2026-05-26; past = end < 2026-05-26
job_rows = [
    ("jobreq-1776184326685", "Ohio Country Fest", "The Ohio Country Fest", "lead",
     "2026-06-15", "2026-06-21", 7, "🔴 FUTURE",
     "ADD CREW NEEDS — 7-day fest, plan needed before show.", "", ""),
    ("jobreq-1775848667250", "Scotty Mccreery", "Chris Stewart - Morris Farms",
     "lead", "2026-05-21", "2026-05-24", 4, "PAST (just)",
     "LEAVE — just past. Note: live invoice work is happening for this client (Sheet 3 🔴).", "", ""),
    ("jobreq-1777325737896", "Miami University Commencement", "Loud&Clear, Inc.",
     "lead", "2026-05-11", "2026-05-17", 7, "PAST",
     "LEAVE — past. Has a sibling row jobreq-1778517711104 with 4 crew_needs — possible duplicate.", "", ""),
    ("jobreq-1775586827221", "Camp Fimfo Concert Series", "JAYSON Entertainment Group",
     "lead", "2026-05-07", "2026-05-09", 3, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1777304800150", "Warrior Conference", "Loud&Clear, Inc.", "booked",
     "2026-04-29", "2026-05-02", 4, "PAST",
     "LEAVE — past, booked.", "", ""),
    ("jobreq-1776229709021", "Luke Combs - Load Out", "Rhino Staging", "lead",
     "2026-04-25", "2026-04-25", 1, "PAST", "LEAVE — past.", "", ""),
    ("jobreq-1776734819124", "Luke Combs - OSU", "Rhino Staging", "booked",
     "2026-04-21", "2026-04-21", 1, "PAST", "LEAVE — past.", "", ""),
    ("jobreq-1775744267941", "Mount St. Joseph - Spring Concert", "Loud&Clear, Inc.",
     "lead", "2026-04-17", "2026-04-18", 2, "PAST",
     "LEAVE — past. (This is the REPOINT TARGET from Sheet 1.)", "", ""),
    ("jobreq-1776106166461", "Storage Unload", "Lighthouse Immersive Cleveland LLC",
     "lead", "2026-04-15", "2026-04-15", 1, "PAST", "LEAVE — past.", "", ""),
    ("jobreq-1775073944709", "Church Concert", "Alive Productions, Inc.", "lead",
     "2026-04-10", "2026-04-10", 1, "PAST", "LEAVE — past.", "", ""),
    ("jobreq-1774997460467", "Pro Football Hall of Fame Enshrinement Week",
     "FEP Live, LLC", "lead", "2026-03-31", "(blank)", 1, "PAST (date wrong)",
     "request_date 2026-03-31 is wrong — the actual event is in August per its quote (recovered-06a99ec9). LEAVE (past status) OR fix the date if Connor wants to re-quote.", "", ""),
]
for i, r in enumerate(job_rows, start=5):
    is_future = "🔴 FUTURE" in str(r[7])
    write_row(ws, i, r, id_cols={1}, decision_col=10, urgent=is_future)
    ws.row_dimensions[i].height = 50

widths = [32, 35, 30, 9, 12, 12, 7, 14, 60, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ════════════════ Sheet 7: Auto-resolved (FYI) ════════════════
ws = wb.create_sheet("7. Auto-resolved (FYI)")
ws["A1"] = "Auto-resolved items — no Connor decision needed"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("These showed up in the fallback audit but prod data is already clean. "
            "Pure code-cleanup once we retire the fallbacks.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 50
ws.merge_cells("A2:D2")

headers = ["Audit item", "Fallback location", "Prod count", "Status"]
write_headers(ws, headers, start_row=4)

auto_rows = [
    ("lunch_minutes dual-write",
     "lib/store/db.ts: timesheetEntryToRow + computeTimeEntry fallback",
     "0 rows where meal_break_1_minutes IS NULL AND lunch_minutes IS NOT NULL",
     "READY to remove dual-write + drop lunch_minutes column."),
    ("rate_card_profile_rows name-match resolver",
     "components/shared/rate-card-editor.tsx + master-rate-card-editor.tsx",
     "0 rows where specialty_id IS NULL",
     "READY to remove resolver code."),
]
for i, r in enumerate(auto_rows, start=5):
    write_row(ws, i, r)
    ws.row_dimensions[i].height = 50

widths = [30, 60, 60, 55]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

wb.save(OUTPUT)
print(f"Wrote: {OUTPUT}")
