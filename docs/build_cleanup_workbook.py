"""Build the cleanup-review workbook for Connor.

One sheet per active-fallback item. Each row has the record GUID (so Connor
can match it to what the 'Show IDs' toggle reveals in the UI), enough
human context to identify the record, a recommendation, and an empty
Decision column for him to fill in.

Run once from the docs/ folder. Output: docs/cleanup-review-2026-05-26.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "C:/amplified/Amplified-AOS/docs/cleanup-review-2026-05-26.xlsx"

# Styling shared across sheets
HEADER_FILL = PatternFill("solid", start_color="2A3F5A")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Calibri", bold=True, size=14)
NOTE_FONT   = Font(name="Calibri", italic=True, size=10, color="555555")
ID_FONT     = Font(name="Consolas", size=10, color="666666")
DECISION_FILL = PatternFill("solid", start_color="FFF7E0")
SECTION_FILL  = PatternFill("solid", start_color="F4F0E6")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def write_headers(ws, headers, start_row=1):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[start_row].height = 30

def write_row(ws, row_idx, values, *, id_cols=None, decision_col=None):
    id_cols = id_cols or set()
    for col_idx, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=v)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = THIN_BORDER
        if col_idx in id_cols:
            c.font = ID_FONT
        if decision_col is not None and col_idx == decision_col:
            c.fill = DECISION_FILL

# ─────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # clean slate, we name our own sheets

# ═════════════════════════ README ════════════════════════════
ws = wb.create_sheet("README")
ws["A1"] = "Cleanup review — fallback retirement"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Date generated: 2026-05-26   |   Env: dev (Supabase ovtbvnfhteqxnyirzctt)"
ws["A3"].font = NOTE_FONT
ws["A5"] = "Purpose"
ws["A5"].font = Font(bold=True, size=12)
ws["A6"] = (
    "An audit of the timekeeping/quote/invoice rewrite found a handful of code "
    "fallback patterns that protect against legacy / unmigrated data. The user "
    "wants to retire those fallbacks. Each fallback corresponds to a set of "
    "records that need a decision (delete, repoint, merge, etc.).\n\n"
    "Sheets in this workbook list the affected records with their record IDs. "
    "Use the 'Show IDs' toggle in the AOS sidebar to reveal the same IDs in "
    "the live UI so you can verify before deciding."
)
ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[6].height = 100

ws["A8"] = "How to use"
ws["A8"].font = Font(bold=True, size=12)
ws["A9"] = (
    "Each sheet has a 'Decision (Connor)' column highlighted in pale yellow. "
    "Fill it with one of: KEEP / DELETE / REPOINT TO <id> / MERGE WITH <id> / "
    "BACKFILL / NEEDS DISCUSSION. Add notes if useful. After you're done, "
    "I'll generate the SQL and run it on dev (then queue the same on prod)."
)
ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[9].height = 80

ws["A11"] = "Sheet index"
ws["A11"].font = Font(bold=True, size=12)
index_rows = [
    ("1. Timesheet stragglers",        "3 timesheets with NULL job_id. Kill the legacy: picker mode."),
    ("2. Quote lines (recovered)",     "12 frozen recovered-* quotes whose lines have NULL specialty_id (191 lines)."),
    ("3. Invoice lines NULL FK",       "42 invoice lines missing position_id and/or specialty_id."),
    ("4. Jobs missing crew planning",  "9 active job_requests with 0 crew_needs."),
    ("5. Auto-resolved (FYI)",         "2 items already clean on dev — no decision needed, just code cleanup."),
]
for i, (name, desc) in enumerate(index_rows, start=12):
    ws.cell(row=i, column=1, value=name).font = Font(bold=True)
    ws.cell(row=i, column=2, value=desc)
    ws.row_dimensions[i].height = 20

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 90

# ════════ Sheet 1: Timesheet stragglers ═════════════════════════
ws = wb.create_sheet("1. Timesheet stragglers")
ws["A1"] = "Cleanup #1 — Timesheets with NULL job_id"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Fallback being retired: the 'Legacy (no Job linked)' picker mode in timekeeping.tsx "
            "and the parallel job_id/jobSheetId branching in timesheet-review.tsx. "
            "Decide per timesheet: delete the whole timesheet (drops entries too), or re-point to a job_request.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 50
ws.merge_cells("A2:H2")

headers = [
    "Timesheet ID", "Job Sheet ID",
    "JS client (free-text)", "JS event", "JS date",
    "Entries", "w/ hours", "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

rows_data = [
    ("timesheet-jobsheet-1774846347019", "jobsheet-1774846347019",
     "Sunbelt Ground Protection Division", "Flooring", "2026-03-28",
     4, 2,
     "DELETE — past, no quote, no approval, no invoice. Only 2 of 4 rows have hours.",
     "", ""),
    ("timesheet-jobsheet-1774846277546", "jobsheet-1774846277546",
     "Sunbelt Ground Protection DIvision (typo)", "Flooring Install", "2026-03-28",
     9, 0,
     "DELETE — empty skeleton + near-duplicate of the row above.",
     "", ""),
    ("timesheet-jobsheet-1774844548762", "jobsheet-1774844548762",
     "Loud and Clear (note: separate clients row from 'Loud&Clear, Inc.')",
     "Spring Concert", "2026-04-17",
     13, 13,
     "REPOINT TO jobreq-1775744267941 (Mount St. Joseph - Spring Concert, Loud&Clear Inc., 2026-04-17). Same date + venue type; very likely the same event captured under the duplicate client name.",
     "", ""),
]
for i, r in enumerate(rows_data, start=5):
    write_row(ws, i, r, id_cols={1, 2}, decision_col=9)
    ws.row_dimensions[i].height = 60

widths = [38, 32, 35, 30, 12, 9, 9, 60, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ════════ Sheet 2: Quote lines (recovered) ═══════════════════════
ws = wb.create_sheet("2. Quote lines (recovered)")
ws["A1"] = "Cleanup #2 — recovered-* quote_lines with NULL specialty_id"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "Fallback being retired: in invoice-draft-editor, quote-detail, quote-pdf-view, "
    "invoice-pdf-view — the '?? legacyPositionMatch ?? l.specialty' chains used to "
    "render frozen recovered quotes whose lines were rebuilt from PDFs.\n\n"
    "These are FROZEN historical records from the Connor incident PDF recovery "
    "(see memory: project_pdf_data_recovery.md). All 191 lines across 12 quotes "
    "have specialty_id=NULL. They display correctly today via the legacy text "
    "fallback. Three paths:\n"
    "  • BACKFILL = disable freeze trigger, name-match specialty_id, re-enable. "
    "Totals unchanged. Removes the fallback need.\n"
    "  • LEAVE = accept that historical records use the legacy display path; "
    "keep just enough fallback for these specific rows.\n"
    "  • DELETE = drop the recovered quote (only if it's no longer needed for "
    "audit / reference). Aggressive."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 150
ws.merge_cells("A2:H2")

headers = [
    "Quote ID", "Event name", "Client",
    "Status", "Lines w/ NULL spec",
    "Total $", "Earliest line date", "Latest line date",
    "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

quote_rows = [
    ("recovered-a65a9683-2026-farm-tour-california", "2026 farm Tour - California",
     "Row Crop LLC - Vaden Group", "issued", 9, 66300, "2026-05-14", "2026-05-16",
     "BACKFILL — sizable $, recent.", "", ""),
    ("recovered-33c5d8eb-ky-event", "KY Event",
     "Loud&Clear, Inc.", "issued", 1, 740, "2026-04-05", "2026-04-05",
     "BACKFILL (or DELETE if duplicates an existing AES_260331_LNC_KYEVENT line).", "", ""),
    ("recovered-271a1ffc-liv-golf-dc", "LIV Golf DC",
     "CSG Productions, LLC.", "signed", 64, 75230, "2026-05-03", "2026-05-12",
     "BACKFILL — largest, signed. Worth a clean FK chain.", "", ""),
    ("recovered-74bb42d2-luke-combs-load-out", "Luke Combs - Load Out",
     "Rhino Staging", "issued", 5, 7648, "2026-04-25", "2026-04-26",
     "BACKFILL.", "", ""),
    ("recovered-6e82573f-miami-university-commencement", "Miami University Commencement",
     "Loud&Clear, Inc.", "signed", 23, 36918, "2026-05-11", "2026-05-17",
     "BACKFILL — signed.", "", ""),
    ("recovered-79529235-mount-st-joseph-spring-concert", "Mount St. Joseph - Spring Concert",
     "Loud&Clear, Inc.", "signed", 6, 4480, "2026-04-17", "2026-04-18",
     "BACKFILL — signed.", "", ""),
    ("recovered-20871778-osu-stadium-load-out", "OSU Stadium - Load Out",
     "Rhino Staging", "issued", 2, 3664, "2026-04-25", "2026-04-25",
     "BACKFILL.", "", ""),
    ("recovered-06a99ec9-pro-football-hall-of-fame-2026-enshrinement-week",
     "Pro Football Hall of Fame 2026 Enshrinement Week",
     "FEP Live, LLC", "signed", 54, 66977, "2026-08-05", "2026-08-10",
     "BACKFILL — signed, sizable, future event.", "", ""),
    ("recovered-ba7ea741-scotty-mccreery", "Scotty Mccreery",
     "Chris Stewart - Morris Farms", "signed", 8, 6665, "2026-05-21", "2026-05-24",
     "BACKFILL — signed, recent.", "", ""),
    ("recovered-b12d4439-the-ohio-country-fest", "The Ohio Country Fest",
     "Aaron Green - Jayson Entertainment Group", "issued", 11, 44592, "2026-06-15", "2026-06-21",
     "BACKFILL — sizable, near-future.", "", ""),
    ("recovered-3d131e98-warrior-conference", "Warrior Conference",
     "Loud&Clear, Inc.", "issued", 4, 4849, "2026-04-29", "2026-05-02",
     "BACKFILL — OR DELETE if duplicates the other Warrior Conference quote below.", "", ""),
    ("recovered-1c3e90df-warrior-conference", "Warrior Conference",
     "Loud&Clear, Inc.", "issued", 4, 4849, "2026-04-29", "2026-05-02",
     "Same as above — TWO copies exist with identical totals. Likely a recovery duplicate. DELETE one?", "", ""),
]
for i, r in enumerate(quote_rows, start=5):
    write_row(ws, i, r, id_cols={1}, decision_col=10)
    ws.row_dimensions[i].height = 55

widths = [52, 50, 35, 12, 10, 12, 15, 15, 60, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ════════ Sheet 3: Invoice lines NULL FK ═══════════════════════
ws = wb.create_sheet("3. Invoice lines NULL FK")
ws["A1"] = "Cleanup #3 — invoice_lines with NULL position_id and/or specialty_id"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "Fallback being retired: the position/specialty text fallback in invoice display + PDF code.\n"
    "42 lines across multiple invoices. Mostly recovered-* invoices from the Connor PDF recovery, "
    "plus a couple of drafts. The legacy 'department' column carries the role name "
    "and can drive a name-match backfill (same pattern as Phase 3's timesheet cleanup).\n"
    "Approach: BACKFILL most by name match. The handful that look like real anomalies "
    "(empty department, zero hours) probably DELETE."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 80
ws.merge_cells("A2:K2")

headers = [
    "Line ID", "Invoice ID", "Invoice no", "Draft",
    "Status", "Type", "Department (legacy)", "Specialty (legacy)",
    "Hours", "Crew", "Total $", "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

invoice_rows = [
    # Drafts
    ("inv-recovered-e6894771-line-1", "inv-recovered-e6894771", "INV-2026-0410-503", True,
     "(draft)", "(unset)", "Stagehand", "", 30.0, 12, 1050,
     "BACKFILL specialty_id by name match on 'Stagehand'.", "", ""),
    ("inv-recovered-e6894771-line-2", "inv-recovered-e6894771", "INV-2026-0410-503", True,
     "(draft)", "(unset)", "Stagehand", "", 30.0, 12, 1050,
     "BACKFILL by name match — exact duplicate of line-1, also flag for review.", "", ""),
    ("inv-1777044736028_2", "inv-1777044736028", "INV-2026-0424-352", True,
     "(draft)", "(unset)", "(empty)", "(empty)", 0, 1, 0,
     "DELETE — zero hours, empty department/specialty. Junk row.", "", ""),
    ("inv-1777044736028_3", "inv-1777044736028", "INV-2026-0424-352", True,
     "(draft)", "(unset)", "(empty)", "(empty)", 0, 1, 0,
     "DELETE — junk row.", "", ""),
    ("il-mozxs1gn-kv2nfzjg", "i-mozxs19h-dkpydfc6", "(draft)", True,
     "(draft)", "final", "", "", 8, 1, 280,
     "Already has specialty_id=spc-01-01; just missing position_id. Derivable via specialty.position_id; CODE FIX rather than data fix.", "", ""),
    ("il-mozxs1gn-exb286hp", "i-mozxs19h-dkpydfc6", "(draft)", True,
     "(draft)", "final", "", "", 10, 1, 650,
     "Already has specialty_id=spc-04-03; CODE FIX (derive position_id via specialty FK).", "", ""),
    # Issued / sent recovered invoices
    ("inv-recovered-58d712e5-line-1", "inv-recovered-58d712e5", "INV-2026-0330-246", False,
     "sent", "final", "Labor", "", 20, 2, 740,
     "BACKFILL — 'Labor' maps to Stagehand (pos-01).", "", ""),
    ("inv-recovered-ff46be24-line-1", "inv-recovered-ff46be24", "INV-2026-0330-627", False,
     "sent", "final", "Labor", "", 15, 3, 540,
     "BACKFILL — 'Labor' → Stagehand (pos-01).", "", ""),
    ("inv-recovered-770ff8c3-line-1", "inv-recovered-770ff8c3", "INV-2026-0419-808", False,
     "superseded", "final", "Operations", "", 7.5, 1, 300,
     "Superseded — LEAVE as historical record OR backfill for cleanliness.", "", ""),
    ("inv-recovered-c8ecc32d-line-1", "inv-recovered-c8ecc32d", "INV-2026-0419-808", False,
     "sent", "final", "Operations", "", 7.5, 1, 300,
     "BACKFILL — 'Operations' → which position? Needs Connor's call.", "", ""),
    ("inv-recovered-c8ecc32d-line-2", "inv-recovered-c8ecc32d", "INV-2026-0419-808", False,
     "sent", "final", "Stagehand", "", 82.5, 11, 3052.5,
     "BACKFILL — 'Stagehand' → pos-01.", "", ""),
    ("inv-recovered-770ff8c3-line-2", "inv-recovered-770ff8c3", "INV-2026-0419-808", False,
     "superseded", "final", "Stagehand", "", 82.5, 11, 3052.5,
     "Superseded copy of above.", "", ""),
    ("inv-recovered-770ff8c3-line-3", "inv-recovered-770ff8c3", "INV-2026-0419-808", False,
     "superseded", "final", "Operations", "", 5, 1, 200, "Superseded.", "", ""),
    ("inv-recovered-c8ecc32d-line-3", "inv-recovered-c8ecc32d", "INV-2026-0419-808", False,
     "sent", "final", "Operations", "", 5, 1, 200, "BACKFILL.", "", ""),
    ("inv-recovered-770ff8c3-line-4", "inv-recovered-770ff8c3", "INV-2026-0419-808", False,
     "superseded", "final", "Fork Op", "", 5, 1, 190, "Superseded; 'Fork Op' → pos-08.", "", ""),
    ("inv-recovered-c8ecc32d-line-4", "inv-recovered-c8ecc32d", "INV-2026-0419-808", False,
     "sent", "final", "Fork Op", "", 5, 1, 190, "BACKFILL — 'Fork Op' → pos-08 Forklift Operator.", "", ""),
    ("inv-recovered-c8ecc32d-line-5", "inv-recovered-c8ecc32d", "INV-2026-0419-808", False,
     "sent", "final", "Stagehand", "", 50, 10, 1850, "BACKFILL — pos-01.", "", ""),
    ("inv-recovered-770ff8c3-line-5", "inv-recovered-770ff8c3", "INV-2026-0419-808", False,
     "superseded", "final", "Stagehand", "", 50, 10, 1850, "Superseded.", "", ""),
    ("inv-recovered-922d482c-line-1", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Operations", "Crew Chief", 15, 1, 600,
     "Superseded; 'Operations / Crew Chief' → needs Connor's call on the right specialty.", "", ""),
    ("inv-recovered-922d482c-line-2", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Stagehand", "Labor", 24, 4, 912,
     "Superseded; 'Stagehand / Labor' → pos-01 + spc-01-01 (or similar).", "", ""),
    ("inv-recovered-922d482c-line-3", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Stagehand", "Labor", 56, 8, 2128, "Superseded.", "", ""),
    ("inv-recovered-922d482c-line-4", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Stagehand", "Labor", 16, 2, 608, "Superseded.", "", ""),
    ("inv-recovered-922d482c-line-5", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Forklift Operator", "Shop", 15, 1, 600,
     "Superseded; 'Forklift Operator / Shop' → pos-08 + spc-08-01.", "", ""),
    ("inv-recovered-922d482c-line-6", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Stagehand", "Labor", 75, 5, 2850, "Superseded.", "", ""),
    ("inv-recovered-922d482c-line-7", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Stagehand", "Labor", 16, 2, 608, "Superseded.", "", ""),
    ("inv-recovered-922d482c-line-8", "inv-recovered-922d482c", "INV-2026-0424-352", False,
     "superseded", "final", "Stagehand", "Labor", 9, 1, 342, "Superseded.", "", ""),
    ("inv-recovered-4a0b36ed-line-1", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Lead", "", 13.5, 1, 540,
     "BACKFILL — 'Lead' → pos-02 Stagehand Lead.", "", ""),
    ("inv-recovered-9601b91a-line-1", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Lead", "", 14.5, 1, 580, "BACKFILL — 'Lead' → pos-02.", "", ""),
    ("inv-recovered-9601b91a-line-2", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Stagehands", "", 110, 11, 4070, "BACKFILL — 'Stagehands' → pos-01.", "", ""),
    ("inv-recovered-4a0b36ed-line-2", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Stagehands", "", 110, 11, 4070, "BACKFILL — pos-01.", "", ""),
    ("inv-recovered-4a0b36ed-line-3", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Stagehands", "", 135.0, 10, 4995, "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-3", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Stagehands", "", 145.0, 10, 5365, "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-4", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Working Fork", "", 13.5, 1, 513,
     "BACKFILL — 'Working Fork' → pos-08 Forklift Operator. Specialty?", "", ""),
    ("inv-recovered-9601b91a-line-4", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Working Fork", "", 14.5, 1, 551, "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-5", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Meal Penalty", "", 42.0, 12, 1470,
     "'Meal Penalty' isn't a position — needs special handling. LEAVE or recategorize as a flat-fee line?", "", ""),
    ("inv-recovered-9601b91a-line-5", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Meal Penalty", "", 42.0, 12, 1470, "Same as above.", "", ""),
    ("inv-recovered-9601b91a-line-6", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Lead", "", 5.5, 1, 220, "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-6", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Lead", "", 5.5, 1, 220, "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-7", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Stagehands", "", 104.5, 19, 3866.5, "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-7", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Stagehands", "", 104.5, 19, 3866.5, "BACKFILL.", "", ""),
    ("inv-recovered-9601b91a-line-8", "inv-recovered-9601b91a", "LC260311", False,
     "sent", "final", "Working Fork", "", 5.5, 1, 209, "BACKFILL.", "", ""),
    ("inv-recovered-4a0b36ed-line-8", "inv-recovered-4a0b36ed", "LC260311", False,
     "sent", "final", "Working Fork", "", 5.5, 1, 209, "BACKFILL.", "", ""),
]
for i, r in enumerate(invoice_rows, start=5):
    write_row(ws, i, r, id_cols={1, 2}, decision_col=13)
    ws.row_dimensions[i].height = 40

widths = [35, 30, 22, 8, 14, 10, 22, 18, 8, 7, 10, 50, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ════════ Sheet 4: Jobs missing crew planning ════════════════════
ws = wb.create_sheet("4. Jobs missing crew planning")
ws["A1"] = "Cleanup #4 — active job_requests with 0 crew_needs"
ws["A1"].font = TITLE_FONT
ws["A2"] = (
    "Fallback being retired: the dashboard's 'legacy understaffed' parallel widget "
    "(components/shared/dashboard.tsx) that reads from job_sheets when crew_needs is empty.\n"
    "9 of 17 active job_requests have 0 crew_needs. For past jobs Connor said 'don't care' "
    "earlier — past jobs can stay empty. Future jobs need crew_needs populated OR the dashboard "
    "fallback dropped."
)
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 70
ws.merge_cells("A2:I2")

headers = [
    "Job request ID", "Job #", "Event", "Client", "Status",
    "Start", "End", "Days", "Crew needs",
    "Past/Future", "Recommendation", "Decision (Connor)", "Notes",
]
write_headers(ws, headers, start_row=4)

# Today is 2026-05-26 per the system prompt — mark past/future accordingly
job_rows = [
    ("jobreq-1779676944207", "AES_26070305_NEW_BIGSHOW", "BigShow", "New Test", "lead",
     "2026-07-03", "2026-07-05", 3, 9, "FUTURE",
     "Has 9 needs already — actually fine. Survey mis-filed it; double-check.", "", ""),
    ("jobreq-1776184326685", "AES_26061521_OCF_OHIOCOUN", "Ohio Country Fest",
     "The Ohio Country Fest", "lead", "2026-06-15", "2026-06-21", 7, 0, "FUTURE",
     "ADD CREW NEEDS — large fest, June 15 start. Needs planning.", "", ""),
    ("jobreq-1775848667250", "AES_26052124_FRM_SCOTTYMC", "Scotty Mccreery",
     "Chris Stewart - Morris Farms", "lead", "2026-05-21", "2026-05-24", 4, 0, "PAST (just)",
     "LEAVE — just-past, work already done.", "", ""),
    ("jobreq-1777899364649", "AES_26051718_TST_TEST", "Test", "Test Client", "lead",
     "2026-05-17", "2026-05-18", 2, 6, "PAST",
     "Test data; has needs. DELETE THE WHOLE JOB once test data is cleared.", "", ""),
    ("jobreq-1777907491509", "AES_26051213_TST_TEST2", "Test2", "Test Client", "quoted",
     "2026-05-12", "2026-05-13", 2, 2, "PAST",
     "Test data. DELETE.", "", ""),
    ("jobreq-1777325737896", "AES_26051117_LNC_MIAMIUNI", "Miami University Commencement",
     "Loud&Clear, Inc.", "lead", "2026-05-11", "2026-05-17", 7, 3, "PAST",
     "Has 3 needs across 7 days — under-populated. LEAVE (past) or backfill from timesheets.", "", ""),
    ("jobreq-1775586827221", "AES_26050709_JAY_CAMPFIMF", "Camp Fimfo Concert Series",
     "JAYSON Entertainment Group", "lead", "2026-05-07", "2026-05-09", 3, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1777304800150", "AES_26042902_LNC_WARRIORC", "Warrior Conference",
     "Loud&Clear, Inc.", "booked", "2026-04-29", "2026-05-02", 4, 0, "PAST",
     "LEAVE — past, booked status.", "", ""),
    ("jobreq-1777429393006", "AES_260429_RHI_RIGINOUT", "Rig IN/OUT",
     "Rhino Staging", "booked", "2026-04-29", "(same day)", 1, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1776229709021", "AES_260425_RHI_LUKECOMB", "Luke Combs - Load Out",
     "Rhino Staging", "lead", "2026-04-25", "(same day)", 1, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1776734819124", "AES_260421_RHI_LUKECOMB", "Luke Combs - OSU",
     "Rhino Staging", "booked", "2026-04-21", "(same day)", 1, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1775744267941", "AES_26041718_LNC_MOUNTSTJ", "Mount St. Joseph - Spring Concert",
     "Loud&Clear, Inc.", "lead", "2026-04-17", "2026-04-18", 2, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1775344443515", "AES_260417_ALV_REVIVALN", "Revival Night",
     "Alive Productions, Inc.", "lead", "2026-04-17", "(same day)", 1, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1776106166461", "AES_260415_LHI_STORAGEU", "Storage Unload",
     "Lighthouse Immersive Cleveland LLC", "lead", "2026-04-15", "(same day)", 1, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1775073944709", "AES_260410_ALV_CHURCHCO", "Church Concert",
     "Alive Productions, Inc.", "lead", "2026-04-10", "(same day)", 1, 0, "PAST",
     "LEAVE — past.", "", ""),
    ("jobreq-1775064575180", "AES_260331_LNC_KYEVENT", "KY Event", "Loud&Clear, Inc.", "lead",
     "2026-03-31", "(same day)", 1, 0, "PAST",
     "LEAVE — past. Also one of the original duplicate clusters (memory entry #2).", "", ""),
    ("jobreq-1774997460467", "AES_260331_FEP_PROFOOTB",
     "Pro Football Hall of Fame 2026 Enshrinement Week", "FEP Live, LLC", "lead",
     "2026-03-31", "(same day)", 1, 0, "PAST",
     "LEAVE — past. (jr.request_date is wrong; event is in August per its quote.)", "", ""),
]
for i, r in enumerate(job_rows, start=5):
    write_row(ws, i, r, id_cols={1}, decision_col=12)
    ws.row_dimensions[i].height = 45

widths = [32, 28, 35, 30, 10, 12, 12, 7, 11, 12, 55, 22, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ════════ Sheet 5: Auto-resolved (FYI) ═══════════════════════════
ws = wb.create_sheet("5. Auto-resolved (FYI)")
ws["A1"] = "Auto-resolved items — no Connor decision needed"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("These showed up in the fallback audit but the underlying data on dev is "
            "already clean. Code-cleanup only. Listed here so we can confirm prod is "
            "the same shape before retiring the fallbacks.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 50
ws.merge_cells("A2:E2")

headers = ["Audit item", "Fallback location", "Dev count", "Prod survey query", "Status"]
write_headers(ws, headers, start_row=4)

auto_rows = [
    ("lunch_minutes dual-write",
     "lib/store/db.ts:timesheetEntryToRow + computeTimeEntry fallback",
     "0 rows where meal_break_1_minutes IS NULL AND lunch_minutes IS NOT NULL",
     "SELECT count(*) FROM timesheet_entries WHERE meal_break_1_minutes IS NULL AND lunch_minutes IS NOT NULL;",
     "READY to remove dual-write + drop lunch_minutes column once prod confirms 0."),
    ("rate_card_profile_rows name-match",
     "components/shared/rate-card-editor.tsx + master-rate-card-editor.tsx (resolveSpecialtyId)",
     "0 rows where specialty_id IS NULL on rate_card_profile_rows",
     "SELECT count(*) FROM rate_card_profile_rows WHERE specialty_id IS NULL;",
     "READY to remove resolver code once prod confirms 0."),
]
for i, r in enumerate(auto_rows, start=5):
    write_row(ws, i, r)
    ws.row_dimensions[i].height = 60

widths = [30, 60, 50, 70, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header rows on each data sheet
for sheet_name in ["1. Timesheet stragglers", "2. Quote lines (recovered)",
                   "3. Invoice lines NULL FK", "4. Jobs missing crew planning",
                   "5. Auto-resolved (FYI)"]:
    wb[sheet_name].freeze_panes = "A5"

wb.save(OUTPUT)
print(f"Wrote: {OUTPUT}")
